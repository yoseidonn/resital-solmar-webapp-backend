from fastapi import APIRouter, HTTPException, UploadFile, File
import os
import openpyxl
import base64
import uuid
from datetime import datetime, date
from schemas import APISReportFileCreate, APISReportFileRead, AdvancedPassengerInformationCreate
from schemas.file_upload import FileUploadRequest, FileUploadResponse
from typing import List
from services import apis_report_file_service, advanced_passenger_service
from models import Villa

router = APIRouter()

UPLOAD_DIR = "media/apis_report_files"
os.makedirs(UPLOAD_DIR, exist_ok=True)

@router.post("/", response_model=FileUploadResponse)
async def create_apis_report_file(request: FileUploadRequest):
    """
    Upload Excel file via base64 content and create individual AdvancedPassengerInformation records
    - Automatically extracts filename and sets upload date
    - Parses Excel and creates database records
    """
    # Decode base64 file content
    try:
        file_content = base64.b64decode(request.file_content)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid base64 file content")
    
    # Generate unique filename using UUID to prevent any collisions
    name_without_ext = os.path.splitext(request.filename)[0]
    extension = os.path.splitext(request.filename)[1]
    unique_id = str(uuid.uuid4())
    timestamp = datetime.now().strftime("%Y%m%d")
    unique_filename = f"{name_without_ext}_{timestamp}_{unique_id}{extension}"
    file_location = os.path.join(UPLOAD_DIR, unique_filename)
    
    # check file extension
    if extension not in ['.xlsx', '.xls']:
        raise HTTPException(status_code=400, detail="Invalid file extension")
        
    # Save file content to disk
    with open(file_location, "wb") as f:
        f.write(file_content)
    
    # Create the file record
    file_data = APISReportFileCreate(
        name=request.filename,  # Keep original filename for display
        date=datetime.now().date(),
        file=file_location
    )
    apis_report_file = await apis_report_file_service.create(file_data)
    
    # Parse Excel and create individual AdvancedPassengerInformation records
    records_created = 0
    error_message = None
    upload_successful = True
    
    try:
        records_created = await _parse_excel_and_create_passenger_info(file_location, apis_report_file.id)
    except Exception as e:
        error_message = f"Failed to parse Excel: {str(e)}"
        upload_successful = False
    
    return FileUploadResponse(
        id=apis_report_file.id,
        filename=request.filename,
        file_path=file_location,
        records_created=records_created,
        upload_successful=upload_successful,
        error=error_message,
        uploaded_at=apis_report_file.uploaded_at.isoformat()
    )

async def _parse_excel_and_create_passenger_info(file_path: str, apis_report_file_id: int) -> int:
    """Parse Excel file and create AdvancedPassengerInformation records"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Get headers from first row
    headers = [cell.value for cell in ws[1]]
    records_created = 0
    errors = 0
    
    print(f"Starting to process Excel file with headers: {headers}")
    
    # Process each data row
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):  # Skip empty rows
            continue
            
        try:
            # Create a dict mapping headers to values
            row_data = dict(zip(headers, row))
            
            # For debugging - print the actual data for problematic rows
            problematic_fields = []
            for field in ['holiday_start_date', 'holiday_end_date', 'date_of_birth', 'foid_issue', 'foid_expiry']:
                if field in row_data and row_data[field] is not None and not isinstance(row_data[field], (date, datetime, str)):
                    problematic_fields.append(f"{field}: {type(row_data[field])} = {row_data[field]}")
            
            if problematic_fields:
                print(f"Row {row_idx} has problematic date fields: {', '.join(problematic_fields)}")
            
            # Find villa by name if exists
            villa_id = None
            if row_data.get('villa_name') or row_data.get('accomodation_name'):
                villa_name = row_data.get('villa_name') or row_data.get('accomodation_name')
                villa = await Villa.filter(villa_name=villa_name).first()
                if villa:
                    villa_id = villa.id
            
            # Parse dates safely
            def parse_date(date_val):
                if date_val is None or date_val == '':
                    return None
                if isinstance(date_val, datetime):
                    return date_val.date()
                elif isinstance(date_val, date):
                    return date_val
                elif isinstance(date_val, str):
                    if not date_val.strip():
                        return None
                    try:
                        return datetime.strptime(date_val, '%Y-%m-%d').date()
                    except:
                        try:
                            return datetime.strptime(date_val, '%d/%m/%Y').date()
                        except:
                            print(f"Failed to parse date from: '{date_val}', type: {type(date_val)}")
                            return None
                print(f"Unhandled date type: {type(date_val)}, value: {date_val}")
                return None
            
            def parse_datetime(datetime_val):
                if datetime_val is None or datetime_val == '':
                    return None
                if isinstance(datetime_val, datetime):
                    return datetime_val
                elif isinstance(datetime_val, str):
                    if not datetime_val.strip():
                        return None
                    try:
                        return datetime.strptime(datetime_val, '%Y-%m-%d %H:%M:%S')
                    except:
                        try:
                            return datetime.strptime(datetime_val, '%d/%m/%Y %H:%M:%S')
                        except:
                            print(f"Failed to parse datetime from: '{datetime_val}', type: {type(datetime_val)}")
                            return None
                print(f"Unhandled datetime type: {type(datetime_val)}, value: {datetime_val}")
                return None
            
            # Create AdvancedPassengerInformation record with robust handling of missing values
            passenger_data = AdvancedPassengerInformationCreate(
                account_name=str(row_data.get('account_name', '') or ''),
                country=str(row_data.get('country', '') or ''),
                passenger_name=str(row_data.get('passenger_name', '') or ''),
                opportunity_name=int(row_data.get('opportunity_name', 0)) if row_data.get('opportunity_name') else 0,
                accomodation_name=str(row_data.get('accomodation_name', '') or ''),
                holiday_start_date=parse_date(row_data.get('holiday_start_date')),
                holiday_end_date=parse_date(row_data.get('holiday_end_date')),
                age=int(row_data.get('age', 0)) if row_data.get('age') else 0,
                date_of_birth=parse_date(row_data.get('date_of_birth')),
                country_of_issue=str(row_data.get('country_of_issue', '') or ''),
                document_type=str(row_data.get('document_type', '') or ''),
                foid_number=str(row_data.get('foid_number', '') or ''),
                foid_issue=parse_datetime(row_data.get('foid_issue')),
                foid_expiry=parse_datetime(row_data.get('foid_expiry')),
                nationality=str(row_data.get('nationality', '') or ''),
                villa_id=villa_id,
                apis_report_file_id=apis_report_file_id
            )
            
            await advanced_passenger_service.create(passenger_data)
            records_created += 1
            
        except Exception as e:
            errors += 1
            print(f"Error processing row {row_idx}: {str(e)}")
            # Print the problematic row data for debugging
            if 'row_data' in locals():
                print(f"Row data: {row_data}")
            continue
    
    wb.close()
    print(f"Finished processing Excel file. Created {records_created} records. Encountered {errors} errors.")
    return records_created

# Manual creation endpoint for empty file records
@router.post("/manual", response_model=APISReportFileRead)
async def create_empty_apis_report_file(name: str, date: str):
    """Create empty APIs report file record without file upload"""
    try:
        file_date = datetime.strptime(date, '%Y-%m-%d').date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Date must be in YYYY-MM-DD format")
    
    file_data = APISReportFileCreate(
        name=name,
        date=file_date,
        file=""  # No physical file
    )
    return await apis_report_file_service.create(file_data)

# Legacy upload endpoint for backward compatibility (multipart form)
@router.post("/upload", response_model=FileUploadResponse)
async def upload_apis_report_file_legacy():
    """Legacy upload endpoint - now deprecated, use POST / with JSON"""
    raise HTTPException(
        status_code=410, 
        detail="This endpoint is deprecated. Use POST / with JSON body containing filename and file_content (base64)"
    )

@router.get("/", response_model=List[APISReportFileRead])
async def list_apis_report_files():
    files = await apis_report_file_service.list_all()
    return files

@router.get("/{file_id}", response_model=APISReportFileRead)
async def get_apis_report_file(file_id: int):
    file = await apis_report_file_service.get_by_id(file_id)
    if not file:
        raise HTTPException(status_code=404, detail="APISReportFile not found")
    return file

@router.put("/{file_id}", response_model=APISReportFileRead)
async def update_apis_report_file(file_id: int, file: APISReportFileCreate):
    obj = await apis_report_file_service.update(file_id, file)
    if not obj:
        raise HTTPException(status_code=404, detail="APISReportFile not found")
    return obj

@router.delete("/{file_id}")
async def delete_apis_report_file(file_id: int):
    deleted = await apis_report_file_service.delete(file_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="APISReportFile not found")
    return {"ok": True} 