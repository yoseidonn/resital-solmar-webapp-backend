from fastapi import APIRouter, HTTPException, UploadFile, File
import os
import openpyxl
import base64
import uuid
from datetime import datetime, date
from schemas import ResortReportFileCreate, ResortReportFileRead, ResortReportCreate
from schemas.file_upload import FileUploadRequest, FileUploadResponse
from typing import List
from services import resort_report_file_service, resort_report_service
from models import Villa

router = APIRouter()

UPLOAD_DIR = "media/resort_report_files"
os.makedirs(UPLOAD_DIR, exist_ok=True)

@router.post("/", response_model=FileUploadResponse)
async def create_resort_report_file(request: FileUploadRequest):
    """
    Upload Excel file via base64 content and create individual ResortReport records
    - Automatically extracts filename and sets upload date
    - Parses Excel and creates database records
    """
    # Decode base64 file content
    try:
        file_content = base64.b64decode(request.file_content)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid base64 file content")
    
    # Generate unique filename using UUID to prevent any collisions
    name_without_ext = os.path.splitext(request.filename)[0]
    extension = os.path.splitext(request.filename)[1]
    unique_id = str(uuid.uuid4())
    timestamp = datetime.now().strftime("%Y%m%d")
    unique_filename = f"{name_without_ext}_{timestamp}_{unique_id}{extension}"
    file_location = os.path.join(UPLOAD_DIR, unique_filename)
    
    # check file extension
    if extension not in ['.xlsx', '.xls']:
        raise HTTPException(status_code=400, detail="Invalid file extension")
    
    # Save file content to disk
    with open(file_location, "wb") as f:
        f.write(file_content)
    
    # Create the file record
    file_data = ResortReportFileCreate(
        name=request.filename,  # Keep original filename for display
        date=datetime.now().date(),
        file=file_location
    )
    resort_report_file = await resort_report_file_service.create(file_data)
    
    # Parse Excel and create individual ResortReport records
    records_created = 0
    error_message = None
    upload_successful = True
    
    try:
        records_created = await _parse_excel_and_create_reports(file_location, resort_report_file.id)
    except Exception as e:
        error_message = f"Failed to parse Excel: {str(e)}"
        upload_successful = False
    
    return FileUploadResponse(
        id=resort_report_file.id,
        filename=request.filename,
        file_path=file_location,
        records_created=records_created,
        upload_successful=upload_successful,
        error=error_message,
        uploaded_at=resort_report_file.uploaded_at.isoformat()
    )

async def _parse_excel_and_create_reports(file_path: str, resort_report_file_id: int) -> int:
    """Parse Excel file and create ResortReport records"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Get headers from first row
    headers = [cell.value for cell in ws[1]]
    records_created = 0
    
    # Process each data row
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):  # Skip empty rows
            continue
            
        try:
            # Create a dict mapping headers to values
            row_data = dict(zip(headers, row))
            
            # Find villa by name if exists
            villa_id = None
            if row_data.get('villa_name') or row_data.get('accomodation_name'):
                villa_name = row_data.get('villa_name') or row_data.get('accomodation_name')
                villa = await Villa.filter(villa_name=villa_name).first()
                if villa:
                    villa_id = villa.id
            
            # Parse dates safely
            def parse_date(date_val):
                if date_val is None or date_val == '':
                    return None
                if isinstance(date_val, datetime):
                    return date_val.date()
                elif isinstance(date_val, date):
                    return date_val
                elif isinstance(date_val, str):
                    if not date_val.strip():
                        return None
                    try:
                        return datetime.strptime(date_val, '%Y-%m-%d').date()
                    except:
                        try:
                            return datetime.strptime(date_val, '%d/%m/%Y').date()
                        except:
                            return None
                return None
            
            def parse_datetime(datetime_val):
                if datetime_val is None or datetime_val == '':
                    return None
                if isinstance(datetime_val, datetime):
                    return datetime_val
                elif isinstance(datetime_val, str):
                    if not datetime_val.strip():
                        return None
                    try:
                        return datetime.strptime(datetime_val, '%Y-%m-%d %H:%M:%S')
                    except:
                        try:
                            return datetime.strptime(datetime_val, '%d/%m/%Y %H:%M:%S')
                        except:
                            return None
                return None
            
            # Create ResortReport record with robust handling of missing values
            report_data = ResortReportCreate(
                accomodation_name=str(row_data.get('accomodation_name', '') or ''),
                villa_id=villa_id,
                supplier=str(row_data.get('supplier', '') or ''),
                resort=str(row_data.get('resort', '') or ''),
                opportunity_name=int(row_data.get('opportunity_name', 0)) if row_data.get('opportunity_name') else 0,
                lead_passenger=str(row_data.get('lead_passenger', '') or ''),
                holiday_start_date=parse_date(row_data.get('holiday_start_date')),
                holiday_end_date=parse_date(row_data.get('holiday_end_date')),
                total_number_of_passenger=int(row_data.get('total_number_of_passenger', 0)) if row_data.get('total_number_of_passenger') else 0,
                adults=int(row_data.get('adults', 0)) if row_data.get('adults') else 0,
                children=int(row_data.get('children', 0)) if row_data.get('children') else 0,
                infants=int(row_data.get('infants', 0)) if row_data.get('infants') else 0,
                flight_arrival_date=parse_date(row_data.get('flight_arrival_date')),
                flight_arrival_time=parse_datetime(row_data.get('flight_arrival_time')),
                depature_date=parse_date(row_data.get('depature_date')),
                departure_flight_time=parse_datetime(row_data.get('departure_flight_time')),
                extras_aggregated=str(row_data.get('extras_aggregated', '') or ''),
                villa_manager_visit_request=str(row_data.get('villa_manager_visit_request', '') or ''),
                live_villa_manager=str(row_data.get('live_villa_manager', '') or ''),
                dt_aff_nane=str(row_data.get('dt_aff_nane', '') or ''),
                resort_report_notes=str(row_data.get('resort_report_notes', '') or ''),
                resort_report_file_id=resort_report_file_id
            )
            
            await resort_report_service.create(report_data)
            records_created += 1
            
        except Exception as e:
            print(f"Error processing row: {e}")
            continue
    
    wb.close()
    return records_created

# Manual creation endpoint for empty file records
@router.post("/manual", response_model=ResortReportFileRead)
async def create_empty_resort_report_file(name: str, date: str):
    """Create empty resort report file record without file upload"""
    try:
        file_date = datetime.strptime(date, '%Y-%m-%d').date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Date must be in YYYY-MM-DD format")
    
    file_data = ResortReportFileCreate(
        name=name,
        date=file_date,
        file=""  # No physical file
    )
    return await resort_report_file_service.create(file_data)

# Legacy upload endpoint for backward compatibility (multipart form)
@router.post("/upload", response_model=FileUploadResponse)
async def upload_resort_report_file_legacy():
    """Legacy upload endpoint - now deprecated, use POST / with JSON"""
    raise HTTPException(
        status_code=410, 
        detail="This endpoint is deprecated. Use POST / with JSON body containing filename and file_content (base64)"
    )

@router.get("/", response_model=List[ResortReportFileRead])
async def list_resort_report_files():
    files = await resort_report_file_service.list_all()
    return files

@router.get("/{file_id}", response_model=ResortReportFileRead)
async def get_resort_report_file(file_id: int):
    file = await resort_report_file_service.get_by_id(file_id)
    if not file:
        raise HTTPException(status_code=404, detail="ResortReportFile not found")
    return file

@router.put("/{file_id}", response_model=ResortReportFileRead)
async def update_resort_report_file(file_id: int, file: ResortReportFileCreate):
    obj = await resort_report_file_service.update(file_id, file)
    if not obj:
        raise HTTPException(status_code=404, detail="ResortReportFile not found")
    return obj

@router.delete("/{file_id}")
async def delete_resort_report_file(file_id: int):
    deleted = await resort_report_file_service.delete(file_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="ResortReportFile not found")
    return {"ok": True} 